#!/usr/bin/env python3
"""
fetch_kalshi_all_categories.py
Fetches ALL Kalshi markets (open + closed + settled + historical) and writes
a 4-sheet Excel workbook identical in structure to the Polymarket output.
"""

import time
import json
import gzip
import os
import re
import requests
from collections import defaultdict
from datetime import date as pydate
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constants ──────────────────────────────────────────────────────────────────
API_BASE    = "https://external-api.kalshi.com/trade-api/v2"
OUTPUT_PATH = r"C:\Users\jarre\OneDrive\Documents\The Projection Room\Kalshi_All_Categories.xlsx"
LOCAL_PATH  = r"C:\Users\jarre\polymarket-film\Kalshi_All_Categories.xlsx"
CACHE_PATH  = r"C:\Users\jarre\polymarket-film\kalshi_markets_cache.json.gz"
KALSHI_URL  = "https://kalshi.com/events/"
CACHE_AGE_H = 6
LIMIT            = 200
HIST_LIMIT       = 1000
MAX_HIST_MARKETS = 1_000_000   # safety cap; date filter (2024+) is the primary stop condition
HIST_MIN_YEAR    = 2024        # only fetch historical markets that closed on/after Jan 1, 2024
DELAY_S          = 0.15

HEADERS = {"Accept": "application/json"}

# Category map: Kalshi API category string -> our label (case-insensitive lookup below)
KALSHI_CAT_MAP = {
    # Exact strings returned by the Kalshi API
    "elections":               "Politics",
    "politics":                "Politics",
    "sports":                  "Sports",
    "entertainment":           "Entertainment",
    "economics":               "Economics",
    "financials":              "Economics",
    "climate and weather":     "Weather",
    "science and technology":  "Science & Tech",
    "crypto":                  "Crypto",
    "companies":               "Business",
    "mentions":                "Media",
    "commodities":             "Economics",
    "world":                   "Geopolitics",
    "social":                  "Culture & Society",
    "health":                  "Health & Medicine",
    "transportation":          "Other",
    # Generic / series-level aliases
    "science":                 "Science & Tech",
    "technology":              "Science & Tech",
    "culture":                 "Culture & Society",
    "weather":                 "Weather",
    "climate":                 "Weather",
    "medicine":                "Health & Medicine",
    "gaming":                  "Gaming",
    "esports":                 "Gaming",
    "geopolitics":             "Geopolitics",
    "law":                     "Law & Legal",
    "legal":                   "Law & Legal",
    "media":                   "Media",
    "news":                    "Media",
    "business":                "Business",
    "awards":                  "Entertainment",
    "movies":                  "Entertainment",
    "tv":                      "Entertainment",
    "music":                   "Entertainment",
}

# Fallback: series_ticker prefix -> our label
SUBCATEGORY_MAP = {
    "KXBTC":    "Crypto",    "KXETH":    "Crypto",    "KXSOL":    "Crypto",
    "KXNQ":     "Economics", "KXSP":     "Economics", "KXINFL":   "Economics",
    "KXFED":    "Economics", "KXGDP":    "Economics", "KXUNEMP":  "Economics",
    "KXPRES":   "Politics",  "KXSEN":    "Politics",  "KXHOUSE":  "Politics",
    "KXGOV":    "Politics",  "KXELEX":   "Politics",
    "KXNFL":    "Sports",    "KXNBA":    "Sports",    "KXMLB":    "Sports",
    "KXNHL":    "Sports",    "KXSOCCER": "Sports",    "KXTENNIS": "Sports",
    "KXMMA":    "Sports",    "KXFORMULA":"Sports",
    "KXOSCAR":  "Entertainment", "KXGOLDEN": "Entertainment",
    "KXEMMY":   "Entertainment", "KXGRAMMY": "Entertainment",
    "KXBOX":    "Gaming",    "KXVG":     "Gaming",
    "KXWX":     "Weather",   "KXHURR":   "Weather",
    "KXPOP":    "Demographics", "KXCOVID": "Health & Medicine",
}

# Title keyword rules -> category
TITLE_RULES = [
    (re.compile(r"\b(oscar|emmy|grammy|golden globe|tony award|film|movie|box office|actor|actress|director|netflix|hulu|disney|hollywood|award|bafta|cannes|sundance)\b", re.I), "Entertainment"),
    (re.compile(r"\b(nfl|nba|mlb|nhl|soccer|mls|tennis|golf|pga|ufc|mma|formula|nascar|olympics|super bowl|world cup|championship|playoff|ncaa)\b", re.I), "Sports"),
    (re.compile(r"\b(bitcoin|ethereum|crypto|btc|eth|sol|doge|xrp|blockchain|defi|nft|web3)\b", re.I), "Crypto"),
    (re.compile(r"\b(president|congress|senate|house rep|election|vote|democrat|republican|gop|trump|biden|harris|primary|ballot|legislation|supreme court)\b", re.I), "Politics"),
    (re.compile(r"\b(fed|federal reserve|interest rate|inflation|gdp|unemployment|recession|cpi|pce|treasury|dollar|s&p|nasdaq|dow|stock)\b", re.I), "Economics"),
    (re.compile(r"\b(weather|hurricane|tornado|earthquake|flood|temperature|snowfall|rainfall|climate|el ni[nn]o)\b", re.I), "Weather"),
    (re.compile(r"\b(ai|artificial intelligence|openai|gpt|llm|robot|space|nasa|fda|drug|vaccine|cancer|covid|pandemic)\b", re.I), "Science & Tech"),
    (re.compile(r"\b(war|nato|ukraine|russia|china|taiwan|israel|iran|nuclear|sanctions|geopolitics|military)\b", re.I), "Geopolitics"),
    (re.compile(r"\b(supreme court|lawsuit|indictment|verdict|trial|legal|prison|criminal|acquit)\b", re.I), "Law & Legal"),
]

ENTERTAINMENT_SUBCATS = {
    re.compile(r"\b(film|movie|cinema|box office|oscar|bafta|cannes|sundance|director|actor|actress|cinemat)\b", re.I): "Film",
    re.compile(r"\b(tv|television|emmy|series|show|streaming|netflix|hulu|disney|hbo|prime video|episode|season)\b", re.I): "TV/Streaming",
    re.compile(r"\b(grammy|music|song|album|artist|singer|band|billboard|spotify)\b", re.I): "Music",
    re.compile(r"\b(golden globe|tony|award|ceremony|winner|nominee)\b", re.I): "Awards",
    re.compile(r"\b(video game|esport|gaming|twitch|steam|xbox|playstation|nintendo)\b", re.I): "Gaming",
}

COLUMNS = [
    "Category", "Subcategory", "Event Title", "Market Question",
    "Market Volume (cts)", "Event Volume (cts)", "24h Volume (cts)",
    "Open Interest (cts)", "Liquidity ($)", "Outcome 1 %",
    "Last Trade Price", "Spread ($)", "Status", "Resolution",
    "Event Start", "Event End", "Market Ticker", "Event Ticker",
    "Series Ticker", "Event URL", "# Outcomes", "Open Markets",
    "Closed Markets", "Settled Markets", "Tags", "Notes",
]

ENTS = {"Entertainment", "Film", "TV/Streaming", "Music", "Awards", "Gaming"}


# ── HTTP helpers ───────────────────────────────────────────────────────────────
SESSION = requests.Session()
SESSION.headers.update(HEADERS)


def _get(path, params=None, retries=4):
    url = f"{API_BASE}{path}"
    for attempt in range(retries):
        try:
            r = SESSION.get(url, params=params, timeout=30)
            if r.status_code == 429:
                wait = 2 ** (attempt + 2)
                print(f"  [rate-limit] sleeping {wait}s", flush=True)
                time.sleep(wait)
                continue
            r.raise_for_status()
            return r.json()
        except requests.RequestException as exc:
            if attempt < retries - 1:
                time.sleep(2 ** (attempt + 1))
            else:
                print(f"  [ERROR] {url} {exc}", flush=True)
                return {}
    return {}


# ── Data helpers ───────────────────────────────────────────────────────────────
def _fp(val):
    """Contract-count from _fp string or number."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _dollars(val):
    """Price in [0,1] range."""
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _date(val):
    if not val:
        return ""
    return str(val)[:10]


# ── Category classification ────────────────────────────────────────────────────
def classify(title, series_category, series_ticker):
    # 1. Kalshi native category
    if series_category:
        cat = KALSHI_CAT_MAP.get(series_category.lower())
        if cat:
            sub = _entertainment_subcat(title) if cat == "Entertainment" else ""
            return cat, sub

    # 2. Series ticker prefix
    for prefix, cat in SUBCATEGORY_MAP.items():
        if series_ticker and series_ticker.upper().startswith(prefix):
            sub = _entertainment_subcat(title) if cat == "Entertainment" else ""
            return cat, sub

    # 3. Title keywords
    for pattern, cat in TITLE_RULES:
        if pattern.search(title or ""):
            sub = _entertainment_subcat(title) if cat == "Entertainment" else ""
            return cat, sub

    return "Other", ""


def _entertainment_subcat(title):
    for pattern, sub in ENTERTAINMENT_SUBCATS.items():
        if pattern.search(title or ""):
            return sub
    return ""


# ── Series map builder ─────────────────────────────────────────────────────────
def fetch_series_map():
    """Returns {series_ticker: {category, title}}"""
    series_map = {}
    cursor = None
    print("[series] Fetching series list...", flush=True)
    while True:
        params = {"limit": 200}
        if cursor:
            params["cursor"] = cursor
        data = _get("/series", params)
        items = data.get("series") or []
        for s in items:
            ticker = s.get("ticker") or s.get("series_ticker") or ""
            series_map[ticker] = {
                "category": s.get("category", ""),
                "title":    s.get("title", ""),
            }
        cursor = data.get("cursor")
        if not cursor or not items:
            break
        time.sleep(DELAY_S)
    print(f"  [series] {len(series_map):,} series loaded", flush=True)
    return series_map


# ── Events fetcher ─────────────────────────────────────────────────────────────
def fetch_events_by_status(status, series_map):
    """Paginate /events?status=... with cursor."""
    all_markets = []
    cursor = None
    total_events = 0
    print(f"\n[events/{status}] fetching...", flush=True)
    while True:
        params = {
            "status": status,
            "with_nested_markets": "true",
            "limit": LIMIT,
        }
        if cursor:
            params["cursor"] = cursor
        data = _get("/events", params)
        events = data.get("events") or []
        for ev in events:
            markets = _extract_markets_from_event(ev, series_map)
            all_markets.extend(markets)
        total_events += len(events)
        print(f"  [{status}] events={total_events:,}  markets={len(all_markets):,}", flush=True)
        cursor = data.get("cursor")
        if not cursor or not events:
            break
        time.sleep(DELAY_S)
    return all_markets


def _extract_markets_from_event(ev, series_map):
    """Flatten one event + its nested markets into row-dicts."""
    rows = []
    ev_ticker  = ev.get("event_ticker", "")
    ser_ticker = ev.get("series_ticker", "") or ""
    ev_title   = ev.get("title", "")
    ev_status  = ev.get("status", "")

    # Use event's own category first, then fall back to series map
    ev_cat_raw = ev.get("category", "")
    ser_info   = series_map.get(ser_ticker, {})
    ser_cat    = ev_cat_raw or ser_info.get("category", "")
    category, subcat = classify(ev_title, ser_cat, ser_ticker)

    markets  = ev.get("markets") or []

    # Derive event dates from first market (events don't always carry date fields)
    if markets:
        ev_start = _date(markets[0].get("open_time") or markets[0].get("created_time"))
        ev_end   = _date(markets[0].get("expected_expiration_time") or markets[0].get("close_time"))
    else:
        ev_start = ev_end = ""

    ev_vol_sum  = 0.0
    ev_24h_sum  = 0.0
    ev_liq_sum  = 0.0
    n_open = n_closed = n_settled = 0

    for m in markets:
        ev_vol_sum += _fp(m.get("volume_fp"))
        ev_24h_sum += _fp(m.get("volume_24h_fp"))
        # Dollar liquidity = top-of-book depth: bid_size*bid_price + ask_size*ask_price
        b_sz = _fp(m.get("yes_bid_size_fp"))
        a_sz = _fp(m.get("yes_ask_size_fp"))
        b_px = float(m.get("yes_bid_dollars") or 0)
        a_px = float(m.get("yes_ask_dollars") or 0)
        ev_liq_sum += b_sz * b_px + a_sz * a_px
        ms = m.get("status", "")
        if ms == "open":      n_open    += 1
        elif ms == "closed":  n_closed  += 1
        elif ms == "settled": n_settled += 1

    for m in markets:
        m_ticker = m.get("ticker", "")
        question = m.get("title", "")
        status_m = m.get("status", "")
        result   = m.get("result", "") or m.get("yes_sub_title", "")

        vol    = _fp(m.get("volume_fp"))
        vol24  = _fp(m.get("volume_24h_fp"))
        oi     = _fp(m.get("open_interest_fp"))

        yes_bid = _dollars(m.get("yes_bid_dollars"))
        yes_ask = _dollars(m.get("yes_ask_dollars"))
        last_p  = _dollars(m.get("last_price_dollars"))
        spread  = round(yes_ask - yes_bid, 4) if yes_ask is not None and yes_bid is not None else None

        yes_bid_pct = round(yes_bid * 100, 1) if yes_bid is not None else None
        last_pct    = round(last_p  * 100, 1) if last_p  is not None else None

        tags   = ", ".join(ev.get("tags") or [])
        ev_url = f"{KALSHI_URL}{ev_ticker}"

        rows.append({
            "Category":            category,
            "Subcategory":         subcat,
            "Event Title":         ev_title,
            "Market Question":     question,
            "Market Volume (cts)": round(vol),
            "Event Volume (cts)":  round(ev_vol_sum),
            "24h Volume (cts)":    round(vol24),
            "Open Interest (cts)": round(oi),
            "Liquidity ($)":       round(ev_liq_sum, 2),
            "Outcome 1 %":         yes_bid_pct,
            "Last Trade Price":    last_pct,
            "Spread ($)":          spread,
            "Status":              status_m or ev_status,
            "Resolution":          result,
            "Event Start":         ev_start,
            "Event End":           ev_end,
            "Market Ticker":       m_ticker,
            "Event Ticker":        ev_ticker,
            "Series Ticker":       ser_ticker,
            "Event URL":           ev_url,
            "# Outcomes":          len(markets),
            "Open Markets":        n_open,
            "Closed Markets":      n_closed,
            "Settled Markets":     n_settled,
            "Tags":                tags,
            "Notes":               "",
        })
    return rows


# ── Historical markets fetcher ─────────────────────────────────────────────────
def fetch_historical(series_map):
    """GET /historical/markets — cursor-based, filtered to HIST_MIN_YEAR+."""
    from datetime import datetime, timezone
    cutoff_ts = int(datetime(HIST_MIN_YEAR, 1, 1, tzinfo=timezone.utc).timestamp())
    cutoff_str = f"{HIST_MIN_YEAR}-01-01"

    all_rows = []
    cursor = None
    total = 0
    skipped_old = 0
    print(f"\n[historical] fetching markets closed >= {HIST_MIN_YEAR}-01-01 ...", flush=True)
    while True:
        params = {"limit": HIST_LIMIT, "min_close_ts": cutoff_ts}
        if cursor:
            params["cursor"] = cursor
        data = _get("/historical/markets", params)
        markets = data.get("markets") or []
        hit_cutoff = False
        for m in markets:
            # Python-side date guard (backup if API filter is partial)
            close = m.get("close_time") or m.get("expiration_time") or ""
            if close and close[:10] < cutoff_str:
                skipped_old += 1
                hit_cutoff = True
                continue
            row = _historical_market_to_row(m, series_map)
            if row:
                all_rows.append(row)
        total += len(markets)
        print(f"  [historical] fetched={total:,}  kept={len(all_rows):,}  skipped_old={skipped_old}", flush=True)
        cursor = data.get("cursor")
        if not cursor or not markets:
            break
        if hit_cutoff:
            print(f"  [historical] Reached pre-{HIST_MIN_YEAR} markets — stopping", flush=True)
            break
        if total >= MAX_HIST_MARKETS:
            print(f"  [historical] Safety cap ({MAX_HIST_MARKETS:,}) reached", flush=True)
            break
        time.sleep(DELAY_S)
    return all_rows


def _historical_market_to_row(m, series_map):
    ev_ticker  = m.get("event_ticker", "")
    ser_ticker = m.get("series_ticker", "") or ""
    m_ticker   = m.get("ticker", "")
    title      = m.get("title", "")
    ev_title   = m.get("event_title", title)
    status_m   = m.get("status", "settled")
    result     = m.get("result", "")

    ser_info   = series_map.get(ser_ticker, {})
    ser_cat    = ser_info.get("category", "")
    category, subcat = classify(ev_title or title, ser_cat, ser_ticker)

    vol  = _fp(m.get("volume_fp"))
    oi   = _fp(m.get("open_interest_fp"))
    # Dollar liquidity from order book depth (notional_value_dollars = $1/contract, not useful)
    b_sz = _fp(m.get("yes_bid_size_fp"))
    a_sz = _fp(m.get("yes_ask_size_fp"))
    b_px = float(m.get("yes_bid_dollars") or 0)
    a_px = float(m.get("yes_ask_dollars") or 0)
    liq  = b_sz * b_px + a_sz * a_px

    yes_bid = _dollars(m.get("yes_bid_dollars"))
    yes_ask = _dollars(m.get("yes_ask_dollars"))
    last_p  = _dollars(m.get("last_price_dollars"))
    spread  = round(yes_ask - yes_bid, 4) if yes_ask is not None and yes_bid is not None else None

    yes_bid_pct = round(yes_bid * 100, 1) if yes_bid is not None else None
    last_pct    = round(last_p  * 100, 1) if last_p  is not None else None

    ev_start = _date(m.get("open_time") or m.get("created_time"))
    ev_end   = _date(m.get("close_time") or m.get("expiration_time"))
    ev_url   = f"{KALSHI_URL}{ev_ticker}" if ev_ticker else ""
    tags     = ", ".join(m.get("tags") or [])

    return {
        "Category":            category,
        "Subcategory":         subcat,
        "Event Title":         ev_title,
        "Market Question":     title,
        "Market Volume (cts)": round(vol),
        "Event Volume (cts)":  round(vol),
        "24h Volume (cts)":    0,
        "Open Interest (cts)": round(oi),
        "Liquidity ($)":       round(liq, 2),
        "Outcome 1 %":         yes_bid_pct,
        "Last Trade Price":    last_pct,
        "Spread ($)":          spread,
        "Status":              status_m,
        "Resolution":          result,
        "Event Start":         ev_start,
        "Event End":           ev_end,
        "Market Ticker":       m_ticker,
        "Event Ticker":        ev_ticker,
        "Series Ticker":       ser_ticker,
        "Event URL":           ev_url,
        "# Outcomes":          1,
        "Open Markets":        0,
        "Closed Markets":      0,
        "Settled Markets":     1,
        "Tags":                tags,
        "Notes":               "historical",
    }


# ── Cache helpers ──────────────────────────────────────────────────────────────
def _load_cache():
    if not os.path.exists(CACHE_PATH):
        return None
    age_h = (time.time() - os.path.getmtime(CACHE_PATH)) / 3600
    if age_h > CACHE_AGE_H:
        return None
    print(f"[cache] Loading from {CACHE_PATH}...", flush=True)
    with gzip.open(CACHE_PATH, "rt", encoding="utf-8") as f:
        return json.load(f)


def _save_cache(rows):
    with gzip.open(CACHE_PATH, "wt", encoding="utf-8") as f:
        json.dump(rows, f)


# ── Excel writer ───────────────────────────────────────────────────────────────
def _header_font():  return Font(name="Arial", bold=True, color="FFFFFF", size=10)
def _body_font():    return Font(name="Arial", size=10)
def _header_fill():  return PatternFill("solid", start_color="2E4057")
def _ent_fill():     return PatternFill("solid", start_color="FFF2CC")
def _blue_font():    return Font(name="Arial", size=10, color="0000FF")


XL_ROW_LIMIT = 1_048_575   # Excel max rows per sheet (minus header)

def build_workbook(all_rows):
    wb = Workbook()

    ent_rows   = [r for r in all_rows if r["Category"] in ENTS or r["Subcategory"] in ENTS]
    all_sorted = sorted(all_rows,  key=lambda r: -r["Market Volume (cts)"])
    ent_sorted = sorted(ent_rows,  key=lambda r: -r["Market Volume (cts)"])

    # Excel has a 1,048,576-row limit per sheet. Cap All Markets at XL_ROW_LIMIT
    # (sorted by volume so highest-activity markets are kept).
    capped = len(all_sorted) > XL_ROW_LIMIT
    display_rows = all_sorted[:XL_ROW_LIMIT]
    if capped:
        print(f"  [excel] All Markets capped at {XL_ROW_LIMIT:,} rows "
              f"(full dataset: {len(all_rows):,})", flush=True)

    cat_totals = defaultdict(lambda: {"rows": 0, "vol": 0, "vol_24h": 0, "oi": 0})
    for r in all_rows:   # summary always uses full dataset
        c = r["Category"]
        cat_totals[c]["rows"]    += 1
        cat_totals[c]["vol"]     += r["Market Volume (cts)"]
        cat_totals[c]["vol_24h"] += r["24h Volume (cts)"]
        cat_totals[c]["oi"]      += r["Open Interest (cts)"]

    summary_rows = sorted(
        [{"Category": c, **v} for c, v in cat_totals.items()],
        key=lambda x: -x["vol"],
    )

    _write_sheet(wb, "All Markets",       COLUMNS, display_rows, highlight_ent=True)
    _write_sheet(wb, "Entertainment",     COLUMNS, ent_sorted,   highlight_ent=False)
    _write_summary(wb, "Category Summary", summary_rows)
    # Raw Data: top 500K rows (avoid double-bloating the file)
    _write_sheet(wb, "Raw Data",          COLUMNS, all_sorted[:500_000], highlight_ent=False, raw=True)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def _write_sheet(wb, name, cols, rows, highlight_ent=False, raw=False):
    ws = wb.create_sheet(name)
    ws.freeze_panes = "A2"

    hf    = _header_font()
    hfill = _header_fill()
    bf    = _body_font()
    blue  = _blue_font()

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = hf
        cell.fill      = hfill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for ri, row in enumerate(rows, 2):
        is_ent = highlight_ent and (row["Category"] in ENTS or row["Subcategory"] in ENTS)
        for ci, col in enumerate(cols, 1):
            val  = row.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = bf
            if is_ent:
                cell.fill = _ent_fill()
            if not raw and col == "Event URL" and val:
                cell.hyperlink = val
                cell.font      = blue
                cell.value     = "Link"

    col_widths = {
        "Category": 18, "Subcategory": 16, "Event Title": 40,
        "Market Question": 50, "Market Volume (cts)": 16,
        "Event Volume (cts)": 16, "24h Volume (cts)": 14,
        "Open Interest (cts)": 16, "Liquidity ($)": 13,
        "Outcome 1 %": 12, "Last Trade Price": 14, "Spread ($)": 10,
        "Status": 10, "Resolution": 12, "Event Start": 12, "Event End": 12,
        "Market Ticker": 22, "Event Ticker": 22, "Series Ticker": 16,
        "Event URL": 10, "# Outcomes": 10, "Open Markets": 12,
        "Closed Markets": 13, "Settled Markets": 14, "Tags": 24, "Notes": 14,
    }
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 14)

    num_fmts = {
        "Market Volume (cts)": "#,##0", "Event Volume (cts)": "#,##0",
        "24h Volume (cts)":    "#,##0", "Open Interest (cts)": "#,##0",
        "Liquidity ($)":       "#,##0.00",
        "Outcome 1 %":         '0.0',
        "Last Trade Price":    '0.0',
        "Spread ($)":          '0.000',
    }
    for ri in range(2, len(rows) + 2):
        for ci, col in enumerate(cols, 1):
            if col in num_fmts:
                ws.cell(row=ri, column=ci).number_format = num_fmts[col]

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.sheet_view.showGridLines = True


def _write_summary(wb, name, summary_rows):
    ws = wb.create_sheet(name)
    ws.freeze_panes = "A2"
    hf    = _header_font()
    hfill = _header_fill()
    bf    = _body_font()

    headers = ["Category", "Market Count", "Total Volume (cts)", "24h Volume (cts)", "Open Interest (cts)"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hf
        cell.fill      = hfill
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(summary_rows, 2):
        vals = [row["Category"], row["rows"], row["vol"], row["vol_24h"], row["oi"]]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = bf
            if ci > 1:
                cell.number_format = "#,##0"

    for ci, w in enumerate([22, 14, 20, 16, 20], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ── Save helper ────────────────────────────────────────────────────────────────
def _save_workbook(wb):
    for path in (OUTPUT_PATH, LOCAL_PATH):
        try:
            wb.save(path)
            print(f"  [save] Saved to {path}", flush=True)
            return path
        except PermissionError:
            print(f"  [save] {path} locked -- trying fallback...", flush=True)
    raise RuntimeError("Could not save to any output path.")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print("=" * 62, flush=True)
    print("  Kalshi -- Full Category Market Fetcher", flush=True)
    print("=" * 62, flush=True)

    all_rows = _load_cache()
    if all_rows:
        print(f"[cache] Loaded {len(all_rows):,} rows from cache", flush=True)
    else:
        # 1. Build series map for category classification
        series_map = fetch_series_map()
        time.sleep(DELAY_S)

        all_rows = []

        # 2-4. Events by status (open, closed, settled)
        for status in ("open", "closed", "settled"):
            rows = fetch_events_by_status(status, series_map)
            all_rows.extend(rows)
            time.sleep(DELAY_S)

        # 5. Historical markets (>~3 months old, flat structure)
        hist_rows = fetch_historical(series_map)
        existing_tickers = {r["Market Ticker"] for r in all_rows}
        added_hist = 0
        for r in hist_rows:
            if r["Market Ticker"] not in existing_tickers:
                all_rows.append(r)
                existing_tickers.add(r["Market Ticker"])
                added_hist += 1
        print(f"[historical] {added_hist:,} new rows after dedup", flush=True)

        print(f"\n[total] {len(all_rows):,} market rows before Excel build", flush=True)
        _save_cache(all_rows)
        print(f"[cache] Saved to {CACHE_PATH}", flush=True)

    print("\n[excel] Building workbook...", flush=True)
    wb = build_workbook(all_rows)

    ent_count = sum(1 for r in all_rows if r["Category"] in ENTS or r["Subcategory"] in ENTS)
    print(f"  All Markets:   {len(all_rows):,} rows", flush=True)
    print(f"  Entertainment: {ent_count:,} rows", flush=True)

    saved = _save_workbook(wb)
    print(f"\n[done] Workbook saved -> {saved}", flush=True)


if __name__ == "__main__":
    main()
