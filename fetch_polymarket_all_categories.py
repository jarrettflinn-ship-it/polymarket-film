#!/usr/bin/env python3
# fetch_polymarket_all_categories.py
# Fetches ALL Polymarket events across every category and exports a 4-sheet Excel file.
# Uses /events endpoint (not /markets) for tag-based category classification.

import time
import re
import json
import sys
import calendar
import gzip
import os
import shutil
from datetime import datetime, date as pydate, timedelta
from collections import defaultdict, Counter

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

API_BASE    = "https://gamma-api.polymarket.com"
OUTPUT_PATH = r"C:\Users\jarre\OneDrive\Documents\The Projection Room\Polymarket_All_Categories.xlsx"
LOCAL_PATH  = r"C:\Users\jarre\polymarket-film\Polymarket_All_Categories.xlsx"
CACHE_PATH  = r"C:\Users\jarre\polymarket-film\polymarket_events_cache.json.gz"
CACHE_AGE_H = 6    # reuse cached events if younger than this many hours
POLY_BASE   = "https://polymarket.com/event/"
LIMIT       = 100
DELAY_S     = 0.5
MAX_RETRIES = 3

C_DARK   = "1A1A2E"
C_DARK2  = "16213E"
C_WHITE  = "FFFFFF"
C_GOLD   = "F0B429"
C_RED    = "E63946"
C_GFILL  = "D4EDDA"
C_STRIPE = "F5F5FF"

COLUMNS = [
    ("Event Title",         "text",   42),
    ("Market Question",     "text",   52),
    ("Outcome 1",           "text",   14),
    ("Outcome 1 %",         "pct",    12),
    ("Outcome 2",           "text",   14),
    ("Outcome 2 %",         "pct",    12),
    ("Market Volume ($)",   "dollar", 18),
    ("Event Volume ($)",    "dollar", 18),
    ("End Date",            "date",   12),
    ("Status",              "text",   10),
    ("Top-Level Category",  "text",   18),
    ("Sub-Category",        "text",   24),
    ("Geo",                 "text",    6),
    ("Tags",                "text",   42),
    ("URL",                 "text",   52),
    ("24h Volume ($)",      "dollar", 16),
    ("1wk Volume ($)",      "dollar", 16),
    ("Open Interest ($)",   "dollar", 16),
    ("Liquidity ($)",       "dollar", 16),
    ("Competitive Score",   "f3",     16),
    ("Comment Count",       "int",    14),
    ("24h Price (pp)",      "f2",     16),
    ("Spread",              "f4",     12),
    ("Last Trade Price",    "f4",     16),
    ("Film Title",          "text",   32),
    ("Content Type",        "text",   14),
]

COL_NAMES  = [c[0] for c in COLUMNS]
COL_FMT    = {c[0]: c[1] for c in COLUMNS}

NUM_FMTS = {
    "dollar": "$#,##0",
    "pct":    "0.0%",
    "date":   "YYYY-MM-DD",
    "f2":     "0.00",
    "f3":     "0.000",
    "f4":     "0.0000",
    "int":    "#,##0",
}

CATEGORY_RULES = [
    ("Entertainment", {
        "movies", "box-office", "oscars", "emmys", "awards", "film",
        "entertainment", "netflix", "top-netflix", "streaming", "pop-culture",
    }),
    ("Politics", {
        "election", "politics", "trump", "biden", "harris", "congress",
        "senate", "president", "democrat", "republican", "us-presidential-election",
        "government", "house-of-representatives", "supreme-court", "political",
    }),
    ("Sports", {
        "nfl", "nba", "mlb", "nhl", "soccer", "tennis", "golf", "mma",
        "ufc", "sports", "fifa", "olympics", "basketball", "baseball",
        "hockey", "football", "superbowl", "champions-league", "epl",
        "premier-league", "formula-1", "f1", "boxing", "rugby", "cricket",
    }),
    ("Crypto", {
        "bitcoin", "crypto", "ethereum", "defi", "nft", "solana", "binance",
        "cryptocurrency", "web3", "blockchain", "btc", "eth", "xrp", "dogecoin",
    }),
    ("Business", {
        "stocks", "finance", "economy", "gdp", "fed", "interest-rate",
        "market", "business", "startups", "earnings", "merger", "ipo", "nasdaq",
    }),
    ("Science", {
        "science", "climate", "space", "nasa", "health", "covid", "ai",
        "technology", "tech", "biotech",
    }),
    ("Culture", {
        "culture", "music", "celebrity", "reality-tv", "grammy", "vma", "television",
    }),
]

SUBCATEGORY_PRIORITY = [
    ("us-presidential-election", "Presidential Election"),
    ("election",         "Elections"),
    ("trump",            "Trump"),
    ("politics",         "Government & Politics"),
    ("nfl",              "NFL"),
    ("nba",              "NBA"),
    ("mlb",              "MLB"),
    ("nhl",              "NHL"),
    ("superbowl",        "Super Bowl"),
    ("champions-league", "Champions League"),
    ("epl",              "Premier League"),
    ("soccer",           "Soccer"),
    ("formula-1",        "Formula 1"),
    ("f1",               "Formula 1"),
    ("ufc",              "UFC/MMA"),
    ("mma",              "UFC/MMA"),
    ("tennis",           "Tennis"),
    ("golf",             "Golf"),
    ("olympics",         "Olympics"),
    ("basketball",       "Basketball"),
    ("baseball",         "Baseball"),
    ("hockey",           "Hockey"),
    ("sports",           "Other Sports"),
    ("box-office",       "Box Office"),
    ("movies",           "Films"),
    ("oscars",           "Oscars"),
    ("emmys",            "Emmys"),
    ("awards",           "Awards"),
    ("netflix",          "Streaming Charts"),
    ("top-netflix",      "Streaming Charts"),
    ("pop-culture",      "Pop Culture"),
    ("entertainment",    "Entertainment"),
    ("bitcoin",          "Bitcoin"),
    ("ethereum",         "Ethereum"),
    ("defi",             "DeFi"),
    ("nft",              "NFT"),
    ("solana",           "Solana"),
    ("crypto",           "Cryptocurrency"),
    ("stocks",           "Equities"),
    ("economy",          "Macroeconomics"),
    ("fed",              "Federal Reserve"),
    ("climate",          "Climate"),
    ("space",            "Space"),
    ("ai",               "Artificial Intelligence"),
    ("health",           "Health"),
    ("covid",            "COVID-19"),
    ("music",            "Music"),
    ("celebrity",        "Celebrity"),
    ("grammy",           "Grammys"),
    ("reality-tv",       "Reality TV"),
]

def safe_float(v):
    try:
        return float(v) if v is not None else None
    except (ValueError, TypeError):
        return None

def safe_int(v):
    try:
        return int(v) if v is not None else None
    except (ValueError, TypeError):
        return None

def parse_date(s):
    if not s:
        return None
    try:
        return datetime.fromisoformat(str(s).replace("Z", "+00:00")).date()
    except Exception:
        pass
    try:
        return pydate.fromisoformat(str(s)[:10])
    except Exception:
        return None

def parse_str_arr(v, fallback=None):
    if isinstance(v, list):
        return v
    if isinstance(v, str):
        try:
            return json.loads(v)
        except Exception:
            pass
    return fallback or []

def classify_category(tag_slugs):
    slug_set = set(tag_slugs)
    for cat, tags in CATEGORY_RULES:
        if slug_set & tags:
            return cat
    return "Other"

def classify_subcategory(tag_slugs):
    slug_set = set(tag_slugs)
    for slug, label in SUBCATEGORY_PRIORITY:
        if slug in slug_set:
            return label
    return "Other"

def extract_film_title(question):
    # Look for title inside ASCII quotes
    m = re.search(r'''["\']([^"\']{2,80}?)["\']''', question)
    if m:
        t = m.group(1).strip()
        if 2 < len(t) < 80:
            return t
    # Pattern: Will X gross/win... where X is the film
    m = re.match(
        r"Will\s+([A-Z][^?]{2,50}?)\s+(?:gross|win|earn|open|debut|make|have|get|be|receive|take)\b",
        question,
    )
    if m:
        t = m.group(1).strip().strip("'\"")
        if len(t) > 2 and t.lower() not in ("it", "this", "the film", "the movie"):
            return t
    return None

def classify_content_type(text):
    t = text.lower()
    if any(w in t for w in ["season", "episode", "series", "emmy", "showrunner",
                             "renewed", "cancelled", "canceled", "finale"]):
        return "TV"
    if any(w in t for w in ["movie", "film", "box office", "opening weekend", "grossing",
                             "oscar", "animated", "sequel", "remake", "director", "actor", "actress"]):
        return "Movie"
    if any(w in t for w in ["album", "song", "artist", "chart", "grammy", "music",
                             "billboard", "tour", "single"]):
        return "Music"
    return "Other"

_session = requests.Session()
_session.headers.update({"Accept": "application/json", "User-Agent": "PolyFilm-Fetcher/1.0"})

def fetch_page(url, params):
    for attempt in range(MAX_RETRIES):
        try:
            r = _session.get(url, params=params, timeout=30)
            r.raise_for_status()
            return r.json()
        except Exception as exc:
            if attempt < MAX_RETRIES - 1:
                wait = 2 ** (attempt + 1)
                print(f"    [retry {attempt+1}/{MAX_RETRIES} in {wait}s] {exc}", flush=True)
                time.sleep(wait)
            else:
                print(f"    [failed after {MAX_RETRIES} retries] {exc}", flush=True)
                return []

def _paginate(label, params_base, seen, all_evs):
    """
    Inner paginator: keep fetching with increasing offset until the batch is
    short (end of results) or we approach the 10,100 API hard cap.
    Returns (events_added, hit_cap, min_end_date) where min_end_date is the
    oldest endDate seen — used by _paginate_range to continue backwards.
    """
    CAP_WARN = 9800   # bail before the 422 kicks in at 10,100
    offset   = 0
    added_total = 0
    hit_cap  = False
    min_date = None
    while True:
        params = dict(params_base, limit=LIMIT, offset=offset,
                      order="end_date_iso", ascending="false")
        batch = fetch_page(f"{API_BASE}/events", params)
        if not isinstance(batch, list) or not batch:
            break
        added = 0
        for ev in batch:
            eid = ev.get("id")
            if eid and eid not in seen:
                seen.add(eid)
                all_evs.append(ev)
                added += 1
                # Track oldest end_date only for newly added events so that
                # _paginate_range can continue from the correct cutoff point.
                # (Duplicates from a prior year-level pass sit at the front of
                # the monthly query results; including them would push min_date
                # to the start of the month and kill the continuation check.)
                ed = str(ev.get("endDate") or ev.get("end_date_iso") or "")[:10]
                if len(ed) == 10:
                    try:
                        ed_d = pydate.fromisoformat(ed)
                        if min_date is None or ed_d < min_date:
                            min_date = ed_d
                    except Exception:
                        pass
        added_total += added
        print(f"  [{label}] offset={offset:6d}  batch={len(batch):3d}  new={added:3d}  grand={len(all_evs):7,}", flush=True)
        if len(batch) < LIMIT:
            break
        offset += LIMIT
        if offset >= CAP_WARN:
            print(f"  [{label}] Approaching API cap — switching to finer chunks", flush=True)
            hit_cap = True
            break
        time.sleep(DELAY_S)
    return added_total, hit_cap, min_date


def _paginate_range(label, start_str, end_str, seen, all_evs):
    """
    Paginate a date-bounded range with automatic continuation.

    When the offset cap is hit, min_date is the oldest endDate among the
    events that were NEWLY added (not already in `seen`).  We restart the
    query with end_date_max=min_date, which slides the window backward and
    picks up the events that were truncated.  Deduplication handles any
    overlap from events that share that exact date.

    prev_min guards against an infinite loop: if two consecutive passes
    return the same min_date we've hit the limit of date-level precision
    (e.g., >9800 events on a single calendar day) and we stop.
    """
    current_end = end_str
    pass_num    = 0
    prev_min    = None

    while True:
        pass_num += 1
        sub_label = label if pass_num == 1 else f"{label}+{pass_num}"
        added, hit_cap, min_date = _paginate(
            sub_label,
            {"closed": "true", "end_date_min": start_str, "end_date_max": current_end},
            seen, all_evs,
        )
        if not hit_cap or min_date is None:
            break
        # Use min_date directly as the new ceiling (no -1 day) so that same-day
        # events truncated by the cap are recovered in the next pass.
        next_end = str(min_date)
        if next_end <= start_str:
            # At or before the start of the range — a same-day query
            # (end_date_min == end_date_max) returns 422; stop here.
            break
        if min_date == prev_min:
            # Two consecutive passes bottomed out at the same date — can't
            # sub-divide further with date-only precision; stop here.
            print(f"  [{label}] Date-precision limit at {min_date} — stopping", flush=True)
            break
        print(f"  [{label}] Cap hit — min_new={min_date}, continuing with end={next_end}", flush=True)
        prev_min     = min_date
        current_end  = next_end
        time.sleep(DELAY_S)


def fetch_all_events(closed):
    """
    Fetch active (closed=False) events with simple offset pagination —
    no cap issue for active markets.
    """
    all_evs, seen = [], set()
    _paginate("active", {"closed": "false"}, seen, all_evs)   # 3-tuple; ignore hit_cap/min_date
    return all_evs


def fetch_closed_chunked():
    """
    Fetch ALL resolved events using year-by-year date-range chunks.

    For years that fit within the 9,800-offset cap a single pass suffices.
    For busier years (election years, crypto peaks, 2026) the function falls
    back to monthly sub-ranges via _paginate_range, which auto-continues if
    even a single month exceeds the cap — no events are ever silently dropped.
    """
    all_evs  = []
    seen     = set()

    this_year = pydate.today().year
    years = list(range(this_year + 2, 2010, -1))

    consecutive_empty = 0

    for year in years:
        y_start = f"{year}-01-01"
        y_end   = f"{year}-12-31"

        added, hit_cap, _ = _paginate(
            str(year),
            {"closed": "true", "end_date_min": y_start, "end_date_max": y_end},
            seen, all_evs,
        )

        if hit_cap:
            # Year too dense — fall back to monthly passes.
            # _paginate_range handles any month that also exceeds the cap by
            # automatically continuing backwards from the cutoff date.
            for month in range(1, 13):
                last_day = calendar.monthrange(year, month)[1]
                m_start  = f"{year}-{month:02d}-01"
                m_end    = f"{year}-{month:02d}-{last_day:02d}"
                _paginate_range(f"{year}-{month:02d}", m_start, m_end, seen, all_evs)
                time.sleep(DELAY_S)

        if added == 0:
            consecutive_empty += 1
            if consecutive_empty >= 3 and year < 2016:
                print(f"  [closed] No events in {year} or prior 2 years — stopping sweep.", flush=True)
                break
        else:
            consecutive_empty = 0

        time.sleep(DELAY_S)

    return all_evs

def process_events(events):
    rows = []
    for ev in events:
        tag_slugs   = [t.get("slug", "") for t in ev.get("tags", [])]
        category    = classify_category(tag_slugs)
        subcategory = classify_subcategory(tag_slugs)
        tags_str    = " | ".join(s for s in tag_slugs if s)
        ev_title    = ev.get("title", "Unknown")
        ev_vol      = safe_float(ev.get("volume"))
        ev_1wk      = safe_float(ev.get("volume1wk"))
        ev_oi       = safe_float(ev.get("openInterest"))
        ev_24h      = safe_float(ev.get("volume24hr"))
        ev_comment  = safe_int(ev.get("commentCount"))
        ev_comp     = safe_float(ev.get("competitive"))
        ev_end      = parse_date(ev.get("endDate"))
        ev_geo      = "Intl" if ev.get("restricted") else "US"
        ev_closed   = bool(ev.get("closed"))
        ev_url      = POLY_BASE + (ev.get("slug") or "")
        markets     = ev.get("markets") or []

        if not markets:
            rows.append({
                "Event Title": ev_title, "Market Question": None,
                "Outcome 1": None, "Outcome 1 %": None,
                "Outcome 2": None, "Outcome 2 %": None,
                "Market Volume ($)": None, "Event Volume ($)": ev_vol,
                "End Date": ev_end, "Status": "Resolved" if ev_closed else "Active",
                "Top-Level Category": category, "Sub-Category": subcategory,
                "Geo": ev_geo, "Tags": tags_str, "URL": ev_url,
                "24h Volume ($)": ev_24h, "1wk Volume ($)": ev_1wk,
                "Open Interest ($)": ev_oi, "Liquidity ($)": None,
                "Competitive Score": ev_comp, "Comment Count": ev_comment,
                "24h Price (pp)": None, "Spread": None, "Last Trade Price": None,
                "Film Title": None, "Content Type": None,
            })
            continue

        for m in markets:
            q        = m.get("question") or ""
            outcomes = parse_str_arr(m.get("outcomes"), ["Yes", "No"])
            prices   = parse_str_arr(m.get("outcomePrices"), ["0.5", "0.5"])
            o1 = outcomes[0] if outcomes else None
            o2 = outcomes[1] if len(outcomes) > 1 else None
            p1 = safe_float(prices[0]) if prices else None
            p2 = safe_float(prices[1]) if len(prices) > 1 else None
            m_vol    = safe_float(m.get("volume"))
            m_1wk    = safe_float(m.get("volume1wk"))
            m_liq    = safe_float(m.get("liquidity"))
            m_delta  = safe_float(m.get("oneDayPriceChange"))
            m_spread = safe_float(m.get("spread"))
            m_ltp    = safe_float(m.get("lastTradePrice"))
            m_closed = bool(m.get("closed"))
            status   = "Resolved" if (m_closed or ev_closed) else "Active"
            combined = (ev_title + " " + q).lower()
            film_title   = extract_film_title(q) if category == "Entertainment" else None
            content_type = classify_content_type(combined) if category == "Entertainment" else None
            rows.append({
                "Event Title": ev_title, "Market Question": q,
                "Outcome 1": o1, "Outcome 1 %": p1,
                "Outcome 2": o2, "Outcome 2 %": p2,
                "Market Volume ($)": m_vol, "Event Volume ($)": ev_vol,
                "End Date": ev_end, "Status": status,
                "Top-Level Category": category, "Sub-Category": subcategory,
                "Geo": ev_geo, "Tags": tags_str, "URL": ev_url,
                "24h Volume ($)": ev_24h,
                "1wk Volume ($)": m_1wk if m_1wk is not None else ev_1wk,
                "Open Interest ($)": ev_oi, "Liquidity ($)": m_liq,
                "Competitive Score": ev_comp, "Comment Count": ev_comment,
                "24h Price (pp)": (m_delta * 100) if m_delta is not None else None,
                "Spread": m_spread, "Last Trade Price": m_ltp,
                "Film Title": film_title, "Content Type": content_type,
            })
    return rows

_HDR_FILL = PatternFill("solid", start_color=C_DARK,  end_color=C_DARK)
_HDR_FONT = Font(bold=True, color=C_WHITE, name="Arial", size=10)
_HDR_ALGN = Alignment(horizontal="center", vertical="center")
_SUB_FILL = PatternFill("solid", start_color=C_DARK2, end_color=C_DARK2)
_SUB_FONT = Font(bold=True, color=C_GOLD,  name="Arial", size=10)
_GRN_FILL = PatternFill("solid", start_color=C_GFILL, end_color=C_GFILL)
_STR_FILL = PatternFill("solid", start_color=C_STRIPE, end_color=C_STRIPE)

def _hdr(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.fill, c.font, c.alignment = _HDR_FILL, _HDR_FONT, _HDR_ALGN
    return c

def _sub(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = _SUB_FILL
    c.font = _SUB_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c

def _write_data_rows(ws, rows_data, start_row=2):
    fmt_seq = [COL_FMT[n] for n in COL_NAMES]
    for i, vals in enumerate(rows_data):
        r = start_row + i
        stripe = (i % 2 == 1)
        for col, (val, fmt) in enumerate(zip(vals, fmt_seq), 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(vertical="top")
            nf = NUM_FMTS.get(fmt)
            if nf:
                c.number_format = nf
            if stripe:
                c.fill = _STR_FILL

def _autowidth(ws, headers, sample_vals, max_w=70):
    ws.freeze_panes = ws.cell(row=2, column=1)
    for i, h in enumerate(headers, 1):
        max_len = len(str(h))
        for rv in sample_vals[:300]:
            v = str(rv[i-1] if rv[i-1] is not None else "")
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, max_w)

def build_raw_data(wb, rows):
    ws = wb.create_sheet("Raw Data")
    ws.row_dimensions[1].height = 18
    for col, name in enumerate(COL_NAMES, 1):
        _hdr(ws, 1, col, name)
    data_vals = [[row.get(n) for n in COL_NAMES] for row in rows]
    _write_data_rows(ws, data_vals)
    _autowidth(ws, COL_NAMES, data_vals)
    print(f"  Sheet 1 (Raw Data):          {len(rows):,} rows", flush=True)

def build_volume_by_category(wb, rows):
    ws = wb.create_sheet("Volume by Category")
    agg        = defaultdict(lambda: {"vol": 0.0, "count": 0})
    cat_totals = defaultdict(lambda: {"vol": 0.0, "count": 0})
    years      = set()
    for row in rows:
        cat = row.get("Top-Level Category", "Other")
        end = row.get("End Date")
        yr  = str(end.year) if isinstance(end, pydate) else "Unknown"
        vol = row.get("Market Volume ($)") or 0.0
        agg[(cat, yr)]["vol"]   += vol
        agg[(cat, yr)]["count"] += 1
        cat_totals[cat]["vol"]  += vol
        cat_totals[cat]["count"] += 1
        years.add(yr)
    sorted_years = sorted((y for y in years if y != "Unknown"), reverse=True)
    if "Unknown" in years:
        sorted_years.append("Unknown")
    cat_order = [c for c, _ in CATEGORY_RULES] + ["Other"]
    headers = ["Category", "Year", "Total Volume ($)", "Market Count", "Avg Vol / Market ($)"]
    for col, h in enumerate(headers, 1):
        _hdr(ws, 1, col, h)
    r = 2
    grand_vol, grand_cnt = 0.0, 0
    for cat in cat_order:
        tot = cat_totals.get(cat)
        if not tot or tot["count"] == 0:
            continue
        grand_vol += tot["vol"]
        grand_cnt += tot["count"]
        c_avg = tot["vol"] / tot["count"] if tot["count"] else 0
        for col, val in enumerate([cat, "All Years", tot["vol"], tot["count"], c_avg], 1):
            c = _sub(ws, r, col, val)
            if col in (3, 5):
                c.number_format = "$#,##0"
            elif col == 4:
                c.number_format = "#,##0"
        r += 1
        for yr in sorted_years:
            d = agg.get((cat, yr))
            if not d or d["count"] == 0:
                continue
            avg = d["vol"] / d["count"] if d["count"] else 0
            for col, val in enumerate([cat, yr, d["vol"], d["count"], avg], 1):
                c = ws.cell(row=r, column=col, value=val)
                c.font = Font(name="Arial", size=9)
                c.alignment = Alignment(horizontal="right" if col > 2 else "left")
                if col == 3:
                    c.number_format = "$#,##0"
                    if d["vol"] > 0:
                        c.fill = _GRN_FILL
                elif col in (4, 5):
                    c.number_format = "#,##0"
            r += 1
        r += 1
    g_avg = grand_vol / grand_cnt if grand_cnt else 0
    for col, val in enumerate(["GRAND TOTAL", "All", grand_vol, grand_cnt, g_avg], 1):
        c = _hdr(ws, r, col, val)
        if col in (3, 5):
            c.number_format = "$#,##0"
        elif col == 4:
            c.number_format = "#,##0"
    ws.freeze_panes = ws.cell(row=2, column=1)
    for col, w in enumerate([22, 10, 20, 14, 24], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    print(f"  Sheet 2 (Volume by Category): built", flush=True)

def build_market_scorecard(wb, rows):
    ws = wb.create_sheet("Market Scorecard")
    cat_order = [c for c, _ in CATEGORY_RULES] + ["Other"]
    cat_data = {cat: {"count": 0, "vol": 0.0, "lrg_q": "", "lrg_v": 0.0,
                      "subcats": defaultdict(float), "dates": []}
                for cat in cat_order}
    for row in rows:
        cat = row.get("Top-Level Category", "Other")
        if cat not in cat_data:
            cat_data[cat] = {"count": 0, "vol": 0.0, "lrg_q": "", "lrg_v": 0.0,
                             "subcats": defaultdict(float), "dates": []}
        vol = row.get("Market Volume ($)") or 0.0
        d = cat_data[cat]
        d["count"] += 1
        d["vol"]   += vol
        q = (row.get("Market Question") or row.get("Event Title") or "")[:80]
        if vol > d["lrg_v"]:
            d["lrg_v"] = vol
            d["lrg_q"] = q
        d["subcats"][row.get("Sub-Category", "Other")] += vol
        end = row.get("End Date")
        if isinstance(end, pydate):
            d["dates"].append(end)
    hdrs = ["Category", "Total Markets", "Total Volume ($)",
            "Largest Market (Question)", "Largest Market Vol ($)",
            "Most Active Sub-Category", "Earliest End Date", "Latest End Date"]
    for col, h in enumerate(hdrs, 1):
        _hdr(ws, 1, col, h)
    r = 2
    for cat in cat_order:
        d = cat_data[cat]
        if d["count"] == 0:
            continue
        top_sub  = max(d["subcats"], key=lambda k: d["subcats"][k]) if d["subcats"] else "N/A"
        dates    = sorted(d["dates"])
        earliest = dates[0] if dates else None
        latest   = dates[-1] if dates else None
        for col, val in enumerate([cat, d["count"], d["vol"], d["lrg_q"], d["lrg_v"],
                                    top_sub, earliest, latest], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="right" if col in (2, 3, 5) else "left", vertical="top")
            if col in (3, 5) and isinstance(val, (int, float)) and val > 0:
                c.number_format = "$#,##0"
                c.fill = _GRN_FILL
            elif col == 2:
                c.number_format = "#,##0"
            elif col in (7, 8) and isinstance(val, pydate):
                c.number_format = "YYYY-MM-DD"
        r += 1
    ws.freeze_panes = ws.cell(row=2, column=1)
    for col, w in enumerate([22, 14, 20, 60, 22, 28, 16, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    print(f"  Sheet 3 (Market Scorecard):  {r-2} category rows", flush=True)

def build_entertainment_extract(wb, rows):
    ws = wb.create_sheet("Entertainment Extract")
    ws.sheet_properties.tabColor = C_RED
    ent = [row for row in rows if row.get("Top-Level Category") == "Entertainment"]
    for col, name in enumerate(COL_NAMES, 1):
        _hdr(ws, 1, col, name)
    data_vals = [[row.get(n) for n in COL_NAMES] for row in ent]
    _write_data_rows(ws, data_vals)
    _autowidth(ws, COL_NAMES, data_vals)
    print(f"  Sheet 4 (Entertainment):     {len(ent):,} rows", flush=True)

def _save_workbook(wb):
    """
    Try saving to the OneDrive path first; fall back to a local path if
    OneDrive has the file locked.  Returns the path actually used.
    """
    for path in (OUTPUT_PATH, LOCAL_PATH):
        try:
            wb.save(path)
            return path
        except PermissionError as exc:
            print(f"  [save] {path} locked ({exc}) — trying next path…", flush=True)
    raise RuntimeError("Could not save to any output path.")


def _load_cache():
    if not os.path.exists(CACHE_PATH):
        return None
    age_h = (time.time() - os.path.getmtime(CACHE_PATH)) / 3600
    if age_h > CACHE_AGE_H:
        print(f"  [cache] Stale ({age_h:.1f}h > {CACHE_AGE_H}h) — refetching.", flush=True)
        return None
    print(f"  [cache] Loading events from cache ({age_h:.1f}h old)…", flush=True)
    with gzip.open(CACHE_PATH, "rt", encoding="utf-8") as f:
        return json.load(f)


def _save_cache(all_evs):
    try:
        with gzip.open(CACHE_PATH, "wt", encoding="utf-8") as f:
            json.dump(all_evs, f)
        size_mb = os.path.getsize(CACHE_PATH) / 1_048_576
        print(f"  [cache] Saved {len(all_evs):,} events to cache ({size_mb:.1f} MB)", flush=True)
    except Exception as exc:
        print(f"  [cache] Warning — could not write cache: {exc}", flush=True)


def main():
    print("=" * 62)
    print("  PolyFilm -- Full Category Market Fetcher")
    print("=" * 62)

    all_evs = _load_cache()

    if all_evs is None:
        print("\n[1/4] Fetching RESOLVED events — year-by-year to bypass 10,100 offset cap...")
        closed_evs = fetch_closed_chunked()
        print(f"      -> {len(closed_evs):,} resolved events")
        print("\n[2/4] Fetching ACTIVE events...")
        active_evs = fetch_all_events(closed=False)
        print(f"      -> {len(active_evs):,} active events")
        seen, all_evs = set(), []
        for ev in active_evs + closed_evs:
            eid = ev.get("id")
            if eid and eid not in seen:
                seen.add(eid)
                all_evs.append(ev)
        print(f"\n      Total unique events: {len(all_evs):,}")
        _save_cache(all_evs)
    else:
        print(f"  [cache] {len(all_evs):,} events loaded — skipping API fetch.")

    print("\n[3/4] Processing events -> market rows...")
    rows = process_events(all_evs)
    print(f"      -> {len(rows):,} market rows")
    print("\n[4/4] Building Excel workbook...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_raw_data(wb, rows)
    build_volume_by_category(wb, rows)
    build_market_scorecard(wb, rows)
    build_entertainment_extract(wb, rows)
    saved_to = _save_workbook(wb)
    print(f"\n  Saved -> {saved_to}")
    if saved_to == LOCAL_PATH:
        print(f"  (OneDrive path was locked; copy to {OUTPUT_PATH} when ready)")
    cat_counts = Counter(r.get("Top-Level Category", "Other") for r in rows)
    total_vol  = sum(r.get("Market Volume ($)") or 0 for r in rows)
    all_dates  = [r.get("End Date") for r in rows if isinstance(r.get("End Date"), pydate)]
    print()
    print("=" * 62)
    print("  SUMMARY")
    print("=" * 62)
    print(f"  Total market rows    {len(rows):>12,}")
    print(f"  Total events         {len(all_evs):>12,}")
    print(f"  Total volume         ${total_vol:>18,.0f}")
    if all_dates:
        print(f"  Date range           {min(all_dates)}  ->  {max(all_dates)}")
    print()
    print(f"  {'Category':<22} {'Markets':>8}  {'Volume':>22}")
    print(f"  {'-'*22}  {'-'*8}  {'-'*22}")
    for cat, _ in CATEGORY_RULES:
        cnt = cat_counts.get(cat, 0)
        if cnt:
            vol = sum(r.get("Market Volume ($)") or 0 for r in rows if r.get("Top-Level Category") == cat)
            print(f"  {cat:<22} {cnt:>8,}  ${vol:>21,.0f}")
    cnt = cat_counts.get("Other", 0)
    if cnt:
        vol = sum(r.get("Market Volume ($)") or 0 for r in rows if r.get("Top-Level Category") == "Other")
        print(f"  {'Other':<22} {cnt:>8,}  ${vol:>21,.0f}")
    print("=" * 62)

if __name__ == "__main__":
    main()
